"""
Build a demo-friendly two-tab spreadsheet from the CustomerGuard health model.
Tab 1 (Summary): portfolio KPIs + a clean, color-coded account overview.
Tab 2 (Full Detail): every signal and sub-score for depth.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from health_model import build_portfolio, portfolio_summary

ARIAL = "Arial"
NAVY = "1F3864"
WHITE = "FFFFFF"
GREEN = "C6EFCE"; GREEN_TXT = "006100"
YELLOW = "FFEB9C"; YELLOW_TXT = "9C6500"
RED = "FFC7CE"; RED_TXT = "9C0006"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

accounts = build_portfolio()
summary = portfolio_summary(accounts)

wb = openpyxl.Workbook()

# ----------------------------------------------------------------------------
# TAB 1: SUMMARY
# ----------------------------------------------------------------------------
ws = wb.active
ws.title = "Summary"

def status_fill(status):
    if status == "Healthy": return PatternFill("solid", fgColor=GREEN), GREEN_TXT
    if status == "At-Risk": return PatternFill("solid", fgColor=YELLOW), YELLOW_TXT
    return PatternFill("solid", fgColor=RED), RED_TXT

# Title
ws["A1"] = "CustomerGuard  |  Customer Health Overview"
ws["A1"].font = Font(name=ARIAL, size=16, bold=True, color=NAVY)
ws["A2"] = f"Security SaaS portfolio  |  as of {summary['as_of']}"
ws["A2"].font = Font(name=ARIAL, size=10, italic=True, color="595959")

# KPI block
kpis = [
    ("Total Accounts", summary["total_accounts"], None),
    ("Total ARR", f"${summary['total_arr']:,}", None),
    ("ARR at Risk", f"${summary['arr_at_risk']:,}", RED),
    ("Healthy", summary["healthy"], GREEN),
    ("At-Risk", summary["at_risk"], YELLOW),
    ("Critical", summary["critical"], RED),
    ("Expansion Signals", summary["expansion_opportunities"], GREEN),
]
row = 4
for i, (label, value, fill) in enumerate(kpis):
    col = 1 + i * 2
    lc = ws.cell(row=row, column=col, value=label)
    lc.font = Font(name=ARIAL, size=9, bold=True, color="595959")
    vc = ws.cell(row=row + 1, column=col, value=value)
    vc.font = Font(name=ARIAL, size=14, bold=True, color=(fill and {GREEN:GREEN_TXT, YELLOW:YELLOW_TXT, RED:RED_TXT}[fill]) or NAVY)

# Account table
head_row = 7
headers = ["Account", "Tier", "ARR", "Health", "Status", "Renewal (days)", "Recommended Action"]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=head_row, column=c, value=h)
    cell.font = Font(name=ARIAL, size=10, bold=True, color=WHITE)
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = border

for r, a in enumerate(accounts, start=head_row + 1):
    fill, txt = status_fill(a["status"])
    values = [
        a["name"], a["tier"], a["arr"], a["health_score"], a["status"],
        a["days_to_renewal"], a["recommended_action"],
    ]
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(name=ARIAL, size=10)
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=(c == 7))
        if c == 3:
            cell.number_format = '$#,##0'
        if c in (4, 5):  # health + status colored
            cell.fill = fill
            cell.font = Font(name=ARIAL, size=10, bold=True, color=txt)
            cell.alignment = Alignment(horizontal="center", vertical="top")

widths = [26, 12, 12, 9, 11, 15, 60]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A8"

# ----------------------------------------------------------------------------
# TAB 2: FULL DETAIL
# ----------------------------------------------------------------------------
ws2 = wb.create_sheet("Full Detail")
ws2["A1"] = "CustomerGuard  |  Full Signal Detail"
ws2["A1"].font = Font(name=ARIAL, size=16, bold=True, color=NAVY)
ws2["A2"] = "All usage, security, and sub-score signals per account"
ws2["A2"].font = Font(name=ARIAL, size=10, italic=True, color="595959")

detail_headers = [
    "Account", "Tier", "ARR", "Seats Licensed", "Seats Active",
    "Logins 30d", "Scans 30d", "Open Critical Vulns", "% Remediated",
    "Support Tickets 30d", "Days to Renewal",
    "Adoption", "Engagement", "Security", "Support",
    "Health Score", "Status", "Flag",
]
dhead = 4
for c, h in enumerate(detail_headers, start=1):
    cell = ws2.cell(row=dhead, column=c, value=h)
    cell.font = Font(name=ARIAL, size=9, bold=True, color=WHITE)
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
    cell.border = border

for r, a in enumerate(accounts, start=dhead + 1):
    s = a["sub_scores"]
    values = [
        a["name"], a["tier"], a["arr"], a["seats_licensed"], a["seats_active"],
        a["logins_last_30d"], a["scans_last_30d"], a["open_critical_vulns"],
        a["pct_remediated"], a["support_tickets_30d"], a["days_to_renewal"],
        s["adoption"], s["engagement"], s["security"], s["support"],
        a["health_score"], a["status"], a["flag"],
    ]
    fill, txt = status_fill(a["status"])
    for c, v in enumerate(values, start=1):
        cell = ws2.cell(row=r, column=c, value=v)
        cell.font = Font(name=ARIAL, size=9)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if c == 3:
            cell.number_format = '$#,##0'
            cell.alignment = Alignment(horizontal="right", vertical="center")
        if c == 1:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        if c in (16, 17):
            cell.fill = fill
            cell.font = Font(name=ARIAL, size=9, bold=True, color=txt)

dwidths = [26, 11, 11, 13, 12, 11, 10, 16, 12, 16, 14, 10, 12, 10, 9, 12, 10, 11]
for i, w in enumerate(dwidths, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "B5"

# Legend / note
note_row = dhead + len(accounts) + 2
ws2.cell(row=note_row, column=1,
         value="Note: Synthetic data. Health = 35% Adoption + 25% Engagement + 25% Security Outcome + 15% Support. Sub-scores and total are 0-100.")
ws2.cell(row=note_row, column=1).font = Font(name=ARIAL, size=9, italic=True, color="595959")

wb.save("CustomerGuard_Health_Demo.xlsx")
print("wrote CustomerGuard_Health_Demo.xlsx")
